"""
report_builder.py
-----------------
Construye el archivo Excel (.xlsx) de reporte de facturas CHEC con 4 hojas:
  1. Resumen    — KPIs generales del portafolio
  2. Facturas   — Tabla maestra con todos los campos extraídos
  3. Por Tipo   — Agrupamiento y subtotales por tipo de factura
  4. Análisis LLM — Análisis generado por IA para cada factura + resumen ejecutivo
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from core.constants import (
    COLOR_VERDE_OSCURO, COLOR_VERDE_MEDIO, COLOR_VERDE_CLARO, COLOR_VERDE_PALIDO,
    COLOR_GRIS_CABECERA, COLOR_BLANCO, COLOR_GRIS_FILA, COLOR_AMARILLO_ALT,
    BORDER_THIN, FORMATO_MONEDA, FORMATO_KW
)
from core.utils import get_nested_val as _get_nested_val

# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _celda_cabecera(cell, texto: str, color_fondo: str = COLOR_VERDE_OSCURO,
                    color_texto: str = COLOR_BLANCO, tamaño: int = 11):
    cell.value = texto
    cell.font = Font(bold=True, color=color_texto, size=tamaño, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=color_fondo)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_THIN


def _celda_dato(cell, valor, alineacion: str = "left", numero_formato: str = None):
    cell.value = valor
    cell.alignment = Alignment(horizontal=alineacion, vertical="center", wrap_text=False)
    cell.border = BORDER_THIN
    if numero_formato:
        cell.number_format = numero_formato


def _autoajustar_columnas(ws, min_width: int = 10, max_width: int = 55):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def _color_por_tipo(tipo: int) -> str:
    colores = {1: COLOR_VERDE_PALIDO, 2: COLOR_VERDE_CLARO,
               3: COLOR_AMARILLO_ALT, 4: "FFE0B2"}
    return colores.get(tipo, COLOR_BLANCO)


# ---------------------------------------------------------------------------
# Hoja 1: Resumen (KPIs)
# ---------------------------------------------------------------------------

def _crear_hoja_resumen(wb: Workbook, facturas: list[dict]):
    ws = wb.create_sheet("📊 Resumen")

    # Título principal
    ws.merge_cells("A1:D1")
    titulo = ws["A1"]
    titulo.value = "REPORTE DE FACTURAS CHEC — RESUMEN EJECUTIVO"
    titulo.font = Font(bold=True, size=16, color=COLOR_BLANCO, name="Calibri")
    titulo.fill = PatternFill("solid", fgColor=COLOR_VERDE_OSCURO)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # KPIs
    total_valor = sum(float(str(_get_nested_val(f, "bloque_control_y_totales.valor_total", 0)).replace(",", ".")) for f in facturas)
    consumos = []
    for f in facturas:
        try:
            val = _get_nested_val(f, "bloque_consumo_energia.kwh_consumidos", 0)
            consumos.append(float(str(val).replace(",", ".")))
        except (ValueError, TypeError):
            consumos.append(0.0)
    total_kwh = sum(consumos)
    municipios = set(_get_nested_val(f, "bloque_datos_cliente.municipio", "N/A") for f in facturas)
    tipos_count = {}
    for f in facturas:
        t = f.get("tipo_factura", 1) # Mantenemos tipo_factura si existe fuera, o default 1
        tipos_count[t] = tipos_count.get(t, 0) + 1

    kpis = [
        ("📋 Total de Facturas Procesadas", len(facturas), None),
        ("💰 Total Facturado", total_valor, FORMATO_MONEDA),
        ("📈 Promedio Valor por Factura", total_valor / max(len(facturas), 1), FORMATO_MONEDA),
        ("⚡ Consumo Total Acumulado", total_kwh, '#,##0.00 "kWh"'),
        ("🏙️ Municipios Únicos", len(municipios), None),
        ("🏠 Facturas Residenciales (Tipo 1+2)", tipos_count.get(1,0)+tipos_count.get(2,0), None),
        ("🏢 Facturas Comerciales (Tipo 3+4)", tipos_count.get(3,0)+tipos_count.get(4,0), None),
    ]

    ws["A3"] = "Indicador"
    ws["B3"] = "Valor"
    _celda_cabecera(ws["A3"], "Indicador", COLOR_VERDE_MEDIO)
    _celda_cabecera(ws["B3"], "Valor", COLOR_VERDE_MEDIO)
    ws.row_dimensions[3].height = 22

    for i, (label, valor, fmt) in enumerate(kpis, 4):
        fill_color = COLOR_GRIS_CABECERA if i % 2 == 0 else COLOR_BLANCO
        fill = PatternFill("solid", fgColor=fill_color)

        c_label = ws.cell(row=i, column=1, value=label)
        c_label.font = Font(bold=True, name="Calibri", size=10)
        c_label.fill = fill
        c_label.border = BORDER_THIN
        c_label.alignment = Alignment(vertical="center")

        c_valor = ws.cell(row=i, column=2, value=valor)
        c_valor.fill = fill
        c_valor.border = BORDER_THIN
        c_valor.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            c_valor.number_format = fmt

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22

    # Detalle por municipio
    row_offset = len(kpis) + 6
    ws.merge_cells(f"A{row_offset}:B{row_offset}")
    ws.cell(row=row_offset, column=1, value="Distribución por Municipio").font = Font(
        bold=True, size=11, color=COLOR_BLANCO, name="Calibri")
    ws.cell(row=row_offset, column=1).fill = PatternFill("solid", fgColor=COLOR_VERDE_MEDIO)
    ws.cell(row=row_offset, column=1).alignment = Alignment(horizontal="center")

    mun_count = {}
    for f in facturas:
        m = f.get("municipio", "N/A")
        mun_count[m] = mun_count.get(m, 0) + 1

    for j, (mun, cnt) in enumerate(sorted(mun_count.items()), row_offset + 1):
        ws.cell(row=j, column=1, value=mun).border = BORDER_THIN
        ws.cell(row=j, column=2, value=cnt).border = BORDER_THIN
        ws.cell(row=j, column=2).alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Hoja 2: Tabla Maestra de Facturas
# ---------------------------------------------------------------------------

COLUMNAS_FACTURAS = [
    ("Archivo",           "archivo_origen",                             "left",   None),
    ("N° Cuenta",         "bloque_control_y_totales.numero_de_cuenta",  "center", None),
    ("Cliente",           "bloque_datos_cliente.nombre",                "left",   None),
    ("Municipio",         "bloque_datos_cliente.municipio",             "left",   None),
    ("Estrato",           "bloque_datos_cliente.estrato",               "center", None),
    ("Período",           "bloque_control_y_totales.factura_del_mes_de", "center", None),
    ("Vence",             "bloque_control_y_totales.fecha_maxima_de_pago", "center", None),
    ("Consumo (kWh)",     "bloque_consumo_energia.kwh_consumidos",       "right",  None),
    ("Valor Total ($)",   "bloque_control_y_totales.valor_total",        "right",  FORMATO_MONEDA),
    ("Lect. Anterior",    "bloque_consumo_energia.lectura_anterior_activa", "center", None),
    ("Lect. Actual",      "bloque_consumo_energia.lectura_actual_activa",   "center", None),
]

COLUMNAS_EXTRA = [
    ("Contribución ($)",  "bloque_consumo_energia.valor_contribucion",  "right",  FORMATO_MONEDA),
    ("Saldo Anterior ($)","bloque_control_y_totales.saldos_meses_anteriores", "right",  FORMATO_MONEDA),
    ("kVAr Penalizados",  "bloque_consumo_energia.consumo_reactiva_en_kvar", "center", None),
]


def _crear_hoja_facturas(wb: Workbook, facturas: list[dict]):
    ws = wb.create_sheet("📄 Facturas")

    todas_cols = COLUMNAS_FACTURAS + COLUMNAS_EXTRA

    # Cabeceras
    for col_idx, (header, _, alin, _fmt) in enumerate(todas_cols, 1):
        _celda_cabecera(ws.cell(row=1, column=col_idx), header)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Filas de datos
    for row_idx, factura in enumerate(facturas, 2):
        tipo = factura.get("tipo_factura", 1)
        fila_color = _color_por_tipo(tipo)
        fill = PatternFill("solid", fgColor=fila_color)

        for col_idx, (_, campo, alin, fmt) in enumerate(todas_cols, 1):
            if "." in campo:
                valor = _get_nested_val(factura, campo)
            else:
                valor = factura.get(campo, "")
                
            if valor == "N/A" or valor is None:
                valor = ""
            c = ws.cell(row=row_idx, column=col_idx)
            _celda_dato(c, valor, alin, fmt)
            c.fill = fill

    _autoajustar_columnas(ws)


# ---------------------------------------------------------------------------
# Hoja 3: Por Tipo (subtotales agrupados)
# ---------------------------------------------------------------------------

def _crear_hoja_por_tipo(wb: Workbook, facturas: list[dict]):
    ws = wb.create_sheet("📈 Por Tipo")

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "ANÁLISIS POR TIPO DE FACTURA"
    t.font = Font(bold=True, size=14, color=COLOR_BLANCO, name="Calibri")
    t.fill = PatternFill("solid", fgColor=COLOR_VERDE_OSCURO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabecera de tabla
    headers = ["Tipo", "Descripción", "Cantidad", "Total Facturado ($)",
               "Promedio ($)", "Consumo Total (kWh)"]
    for col_idx, h in enumerate(headers, 1):
        _celda_cabecera(ws.cell(row=2, column=col_idx), h, COLOR_VERDE_MEDIO)
    ws.row_dimensions[2].height = 22

    tipo_info = {
        1: "Residencial Clásico",
        2: "Residencial Moderno",
        3: "Comercial Alto Consumo",
        4: "Comercial Técnico",
    }

    # Agrupar
    grupos: dict[int, list] = {}
    for f in facturas:
        t_num = f.get("tipo_factura", 0)
        grupos.setdefault(t_num, []).append(f)

    row = 3
    for tipo_num in sorted(grupos.keys()):
        grupo = grupos[tipo_num]
        total = sum(float(str(_get_nested_val(f, "bloque_control_y_totales.valor_total", 0)).replace(",", ".")) for f in grupo)
        kwhs = []
        for f in grupo:
            try:
                val = _get_nested_val(f, "bloque_consumo_energia.kwh_consumidos", 0)
                kwhs.append(float(str(val).replace(",", ".")))
            except (ValueError, TypeError):
                kwhs.append(0.0)
        total_kwh = sum(kwhs)
        promedio = total / max(len(grupo), 1)
        fill = PatternFill("solid", fgColor=_color_por_tipo(tipo_num))

        valores = [tipo_num, tipo_info.get(tipo_num, "Desconocido"),
                   len(grupo), total, promedio, total_kwh]
        formatos = [None, None, None, FORMATO_MONEDA, FORMATO_MONEDA, '#,##0.00']
        alineaciones = ["center", "left", "center", "right", "right", "right"]

        for col_idx, (val, fmt, alin) in enumerate(zip(valores, formatos, alineaciones), 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill = fill
            c.border = BORDER_THIN
            c.alignment = Alignment(horizontal=alin, vertical="center")
            if fmt:
                c.number_format = fmt
        row += 1

    _autoajustar_columnas(ws)


# ---------------------------------------------------------------------------
# Hoja 4: Análisis LLM
# ---------------------------------------------------------------------------

def _crear_hoja_llm(wb: Workbook, analisis_lista: list[dict], resumen_ejecutivo: str):
    ws = wb.create_sheet("🤖 Análisis LLM")

    # Título
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "ANÁLISIS INTELIGENTE DE FACTURAS — GENERADO POR IA (Gemini)"
    t.font = Font(bold=True, size=14, color=COLOR_BLANCO, name="Calibri")
    t.fill = PatternFill("solid", fgColor=COLOR_VERDE_OSCURO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Resumen ejecutivo
    ws.merge_cells("A2:D2")
    ws["A2"].value = "📋 RESUMEN EJECUTIVO DEL PORTAFOLIO"
    ws["A2"].font = Font(bold=True, size=11, color=COLOR_BLANCO, name="Calibri")
    ws["A2"].fill = PatternFill("solid", fgColor=COLOR_VERDE_MEDIO)

    ws.merge_cells("A3:D3")
    c_resumen = ws["A3"]
    c_resumen.value = resumen_ejecutivo
    c_resumen.alignment = Alignment(wrap_text=True, vertical="top")
    c_resumen.border = BORDER_THIN
    c_resumen.fill = PatternFill("solid", fgColor=COLOR_VERDE_PALIDO)
    ws.row_dimensions[3].height = 120

    # Espacio separador
    ws.row_dimensions[4].height = 10

    # Cabecera de análisis individuales
    cab_cols = ["Archivo", "Cliente", "Tipo", "Análisis Detallado"]
    for col_idx, h in enumerate(cab_cols, 1):
        _celda_cabecera(ws.cell(row=5, column=col_idx), h, COLOR_VERDE_MEDIO)
    ws.row_dimensions[5].height = 22

    # Filas de análisis
    for i, item in enumerate(analisis_lista, 6):
        tipo_num = item.get("tipo", 1)
        if not isinstance(tipo_num, int):
            try:
                tipo_num = int(tipo_num)
            except (ValueError, TypeError):
                tipo_num = 1
        fill = PatternFill("solid", fgColor=_color_por_tipo(tipo_num))

        ws.cell(row=i, column=1, value=item.get("archivo", "")).border = BORDER_THIN
        ws.cell(row=i, column=1).fill = fill
        ws.cell(row=i, column=1).alignment = Alignment(vertical="top")

        ws.cell(row=i, column=2, value=item.get("cliente", "")).border = BORDER_THIN
        ws.cell(row=i, column=2).fill = fill
        ws.cell(row=i, column=2).alignment = Alignment(vertical="top")

        ws.cell(row=i, column=3, value=item.get("tipo", 1)).border = BORDER_THIN
        ws.cell(row=i, column=3).fill = fill
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="center", vertical="top")

        c_anal = ws.cell(row=i, column=4, value=item.get("analisis", ""))
        c_anal.alignment = Alignment(wrap_text=True, vertical="top")
        c_anal.border = BORDER_THIN
        c_anal.fill = fill
        ws.row_dimensions[i].height = 130

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 90


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------

def construir_reporte(
    facturas: list[dict],
    analisis_llm: list[dict],
    resumen_ejecutivo: str,
    ruta_salida: str,
) -> str:
    """
    Construye el archivo Excel completo con las 4 hojas.

    Args:
        facturas: Lista de dicts extraídos por extractor.py
        analisis_llm: Lista de análisis individuales del LLM
        resumen_ejecutivo: Texto del resumen ejecutivo del LLM
        ruta_salida: Ruta donde guardar el archivo .xlsx

    Returns:
        Ruta del archivo guardado.
    """
    print("📊 Construyendo reporte Excel...\n")
    wb = Workbook()
    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _crear_hoja_resumen(wb, facturas)
    _crear_hoja_facturas(wb, facturas)
    _crear_hoja_por_tipo(wb, facturas)
    _crear_hoja_llm(wb, analisis_llm, resumen_ejecutivo)

    wb.save(ruta_salida)
    print(f"✅ Reporte guardado en: {ruta_salida}\n")
    return ruta_salida
